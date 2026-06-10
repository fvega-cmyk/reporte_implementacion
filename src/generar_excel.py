"""
Genera el Excel del reporte con hojas 'Resumen' y 'BBDD'.

Cambios respecto a versión anterior:
- Resumen: ahora pivotea por MATERIAL (columnas dinámicas), con texto del PROCESO
  en cada celda. Si el proceso es Rechazado, agrega el detalle entre paréntesis.
  Una fila por sala. Sin fila de total general.
- BBDD: solo las columnas seleccionadas, no toda la base.
"""
from io import BytesIO
from collections import OrderedDict

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config import IDX
from utils import fmt, foto_url


# Columnas que aparecen en la hoja BBDD, en este orden.
# Cada tupla es (encabezado_a_mostrar, llave_IDX, formato_especial)
# formato_especial: None | "fecha" | "foto-N"
COLUMNAS_BBDD = [
    ("CAMPAÑA",         "CAMPANA",        None),
    ("CADENA",          "CADENA",         None),
    ("CÓD.",            "COD",            None),
    ("NOMBRE SALA",     "NOMBRE_SALA",    None),
    ("DIRECCIÓN",       "DIRECCION",      None),
    ("COMUNA",          "COMUNA",         None),
    ("REGION",          "REGION",         None),
    ("MATERIAL",        "MATERIAL",       None),
    ("CANTIDAD",        "CANTIDAD",       None),
    ("FECHA DE ENTREGA","FECHA_ENTREGA",  "fecha"),
    ("REAGENDA",        "REAGENDA",       None),
    ("FECHA REAGENDA",  "FECHA_REAGENDA", "fecha"),
    ("PROCESO",         "PROCESO",        None),
    ("DETALLE",         "DETALLE",        None),
    ("OBSERVACIONES",   "OBSERVACIONES",  None),
    ("FOTO 1",          "FOTO1",          "foto-1"),
    ("FOTO 2",          "FOTO2",          "foto-2"),
    ("FOTO 3",          "FOTO3",          "foto-3"),
    ("FOTO 4",          "FOTO4",          "foto-4"),
]


# Estilos reutilizables
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="1F3864")
ALT_FILL = PatternFill("solid", fgColor="EEF2F9")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
THIN = Side(border_style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
LINK_FONT = Font(name="Calibri", size=11, color="0563C1", underline="single")


def _aplicar_estilos_encabezado(ws, ncols, row=1):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER


def _formatear_proceso(proceso, detalle):
    """
    Devuelve el texto a mostrar en la celda:
    - Si proceso es 'Rechazado' y hay detalle: 'Rechazado (detalle)'
    - Cualquier otro caso: solo el proceso (sin detalle).
    """
    proc = (proceso or "").strip()
    if not proc:
        return ""
    det = (detalle or "").strip()
    if proc.lower() == "rechazado" and det:
        return f"{proc} ({det})"
    return proc


def construir_resumen(ws, filas, campana):
    """
    Construye la hoja Resumen como tabla pivote:
    filas = salas, columnas = materiales (dinámicos), valores = PROCESO (texto).
    """
    # ---- Encabezado superior ----
    ws["A1"] = f"Reporte Diario de Implementación"
    ws["A1"].font = Font(size=14, bold=True, color="1F3864")
    ws.merge_cells("A1:F1")

    ws["A3"] = "Campaña:"
    ws["B3"] = campana
    ws["A3"].font = Font(bold=True)

    ws["A4"] = "Fecha inicio:"
    ws["B4"] = fmt(filas[0][IDX["FECHA_INICIO"]])
    ws["A4"].font = Font(bold=True)

    ws["A5"] = "Fecha término:"
    ws["B5"] = fmt(filas[0][IDX["FECHA_TERMINO"]])
    ws["A5"].font = Font(bold=True)

    # ---- Recolectar materiales únicos en orden de aparición ----
    materiales_orden = OrderedDict()
    for row in filas:
        mat = (row[IDX["MATERIAL"]] or "").strip()
        if mat and mat not in materiales_orden:
            materiales_orden[mat] = True
    materiales = list(materiales_orden.keys())

    # ---- Recolectar datos por sala ----
    # salas[cod] = {
    #   "cod": ..., "nombre": ..., "comuna": ..., "fecha_entrega": ...,
    #   "procesos_por_material": {mat: "texto a mostrar"}
    # }
    salas = OrderedDict()
    duplicados = []
    for row in filas:
        cod = (row[IDX["COD"]] or "").strip()
        if not cod:
            cod = "(sin código)"
        if cod not in salas:
            salas[cod] = {
                "cod": cod,
                "nombre": row[IDX["NOMBRE_SALA"]] or "",
                "comuna": row[IDX["COMUNA"]] or "",
                "fecha_entrega": "",
                "procesos_por_material": {},
            }
        sala = salas[cod]
        # Tomar la primera fecha_entrega no vacía
        if not sala["fecha_entrega"] and row[IDX["FECHA_ENTREGA"]]:
            sala["fecha_entrega"] = fmt(row[IDX["FECHA_ENTREGA"]])

        material = (row[IDX["MATERIAL"]] or "").strip()
        if not material:
            continue
        texto = _formatear_proceso(row[IDX["PROCESO"]], row[IDX["DETALLE"]])

        if material in sala["procesos_por_material"]:
            # Duplicado: dejar el primero y avisar
            duplicados.append((cod, sala["nombre"], material))
        else:
            sala["procesos_por_material"][material] = texto

    if duplicados:
        print(f"  [WARN] {len(duplicados)} duplicados de sala+material en '{campana}' (se conservó el primer valor):")
        for cod, nombre, mat in duplicados[:5]:
            print(f"          - {cod} ({nombre}) / {mat}")
        if len(duplicados) > 5:
            print(f"          ... y {len(duplicados) - 5} más")

    # ---- Encabezados de la tabla ----
    headers = ["Código", "Nombre Sala", "Comuna", "Fecha Entrega"] + materiales
    ncols = len(headers)
    fila_header = 7
    for c, h in enumerate(headers, start=1):
        ws.cell(row=fila_header, column=c, value=h)
    _aplicar_estilos_encabezado(ws, ncols, row=fila_header)

    # ---- Volcar las filas ----
    fila_actual = fila_header + 1
    for cod in sorted(salas.keys()):
        s = salas[cod]
        ws.cell(row=fila_actual, column=1, value=s["cod"]).alignment = CENTER
        ws.cell(row=fila_actual, column=2, value=s["nombre"]).alignment = LEFT
        ws.cell(row=fila_actual, column=3, value=s["comuna"]).alignment = LEFT
        ws.cell(row=fila_actual, column=4, value=s["fecha_entrega"]).alignment = CENTER

        for i, material in enumerate(materiales, start=5):
            valor = s["procesos_por_material"].get(material, "")
            cell = ws.cell(row=fila_actual, column=i, value=valor)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Zebra striping
        if (fila_actual - fila_header) % 2 == 0:
            for c in range(1, ncols + 1):
                ws.cell(row=fila_actual, column=c).fill = ALT_FILL
        for c in range(1, ncols + 1):
            ws.cell(row=fila_actual, column=c).border = BORDER
        fila_actual += 1

    # ---- Anchos de columnas ----
    ws.column_dimensions["A"].width = 12   # Código
    ws.column_dimensions["B"].width = 32   # Nombre Sala
    ws.column_dimensions["C"].width = 18   # Comuna
    ws.column_dimensions["D"].width = 14   # Fecha Entrega
    # Columnas de materiales: ancho según el nombre (con un mínimo de 20 y máximo de 40)
    for i, mat in enumerate(materiales, start=5):
        ancho = max(20, min(40, len(mat) + 2))
        ws.column_dimensions[get_column_letter(i)].width = ancho

    ws.row_dimensions[fila_header].height = 40
    ws.freeze_panes = ws.cell(row=fila_header + 1, column=1).coordinate


def construir_bbdd(ws, filas):
    """
    Construye la hoja BBDD con SOLO las columnas seleccionadas en COLUMNAS_BBDD.
    """
    # Encabezados
    headers = [h for h, _, _ in COLUMNAS_BBDD]
    ws.append(headers)
    _aplicar_estilos_encabezado(ws, len(headers), row=1)

    # Filas
    for r_off, row in enumerate(filas, start=2):
        for c, (header_txt, idx_key, formato) in enumerate(COLUMNAS_BBDD, start=1):
            idx = IDX[idx_key]
            valor_crudo = row[idx] if idx < len(row) else ""

            if formato == "fecha":
                cell = ws.cell(row=r_off, column=c, value=fmt(valor_crudo))
                cell.alignment = CENTER
            elif formato and formato.startswith("foto-"):
                num = formato.split("-")[1]
                ruta = (valor_crudo or "").strip()
                if ruta:
                    url = foto_url(ruta)
                    cell = ws.cell(row=r_off, column=c, value=f"Ver FOTO {num}")
                    cell.hyperlink = url
                    cell.font = LINK_FONT
                    cell.alignment = CENTER
                else:
                    ws.cell(row=r_off, column=c, value="")
            else:
                valor = valor_crudo if valor_crudo not in (None,) else ""
                cell = ws.cell(row=r_off, column=c, value=valor)
                cell.alignment = LEFT

        # Zebra striping
        if r_off % 2 == 0:
            for c in range(1, len(headers) + 1):
                actual = ws.cell(row=r_off, column=c).fill
                if actual.fgColor.rgb in (None, "00000000"):
                    ws.cell(row=r_off, column=c).fill = ALT_FILL

    # Anchos razonables por tipo de columna
    anchos_default = {
        "CAMPAÑA": 28, "CADENA": 14, "CÓD.": 10, "NOMBRE SALA": 32,
        "DIRECCIÓN": 30, "COMUNA": 18, "REGION": 18,
        "MATERIAL": 28, "CANTIDAD": 10, "FECHA DE ENTREGA": 14,
        "REAGENDA": 12, "FECHA REAGENDA": 14, "PROCESO": 14,
        "DETALLE": 35, "OBSERVACIONES": 35,
        "FOTO 1": 12, "FOTO 2": 12, "FOTO 3": 12, "FOTO 4": 12,
    }
    for c, header in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(c)].width = anchos_default.get(header, 18)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(filas) + 1}"


def generar_excel(campana, filas, headers, hoy):
    """
    Construye el Excel completo y devuelve los bytes.
    El parámetro 'headers' se mantiene por compatibilidad con main.py pero no se usa
    (las columnas de BBDD ahora son fijas, definidas en COLUMNAS_BBDD).
    """
    wb = Workbook()
    ws_resumen = wb.active
    ws_resumen.title = "Resumen"
    construir_resumen(ws_resumen, filas, campana)

    ws_bbdd = wb.create_sheet("BBDD")
    construir_bbdd(ws_bbdd, filas)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
